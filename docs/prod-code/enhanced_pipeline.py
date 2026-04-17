"""
Enhanced Pipeline Module - Production Patterns Implementation
Implements enterprise-grade patterns from production code for healthcare claims reporting.
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
from logger import setup_logger, log_section_start, log_section_end, log_error_with_context


def process_by_segments(df, segment_column, segments_dict, logger=None):
    """
    Process data by multiple segments and return aggregated results.
    
    Args:
        df: Input DataFrame
        segment_column: Column name to filter on (e.g., 'SegmentCode', 'LOB', 'BOB')
        segments_dict: Dict mapping segment codes to names
                      Example: {'WGS': 'WGS Market', 'MED': 'Medicaid'}
        logger: Logger instance
    
    Returns:
        Dict of DataFrames, one per segment
    """
    if logger:
        log_section_start(logger, "Multi-Segment Processing")
    
    segment_data = {}
    
    for code, name in segments_dict.items():
        segment_df = df[df[segment_column] == code].copy()
        segment_data[code] = segment_df
        
        if logger:
            logger.info(f"Segment '{name}' ({code}): {len(segment_df)} records")
    
    if logger:
        log_section_end(logger, "Multi-Segment Processing", success=True)
    
    return segment_data


def calculate_advanced_metrics(df, metric_columns, logger=None):
    """
    Calculate comprehensive metrics including totals, rates, and derived values.
    Production pattern: Calculates AA rates, manual claims, 1st/2nd pass processing.
    
    Args:
        df: Input DataFrame with claims data
        metric_columns: List of columns to aggregate
                       Example: ['ITS_RCVD', 'ITS_FNLZD', 'ITS_AA', 'CON_RCVD', 'CON_FNLZD', 'CON_AA']
        logger: Logger instance
    
    Returns:
        DataFrame with calculated metrics
    """
    if logger:
        logger.info("Calculating advanced metrics...")
    
    # Convert columns to numeric
    df_numeric = df.copy()
    for col in metric_columns:
        if col in df_numeric.columns:
            df_numeric[col] = pd.to_numeric(df_numeric[col], errors='coerce')
    
    metrics = {}
    
    # Sum specified columns
    for col in metric_columns:
        if col in df_numeric.columns:
            metrics[f'{col}_Total'] = df_numeric[col].sum()
    
    # Calculate derived metrics (Production pattern)
    # Total Claims = ITS_FNLZD + CON_FNLZD
    metrics['Total_Claims'] = (
        metrics.get('ITS_FNLZD_Total', 0) + 
        metrics.get('CON_FNLZD_Total', 0)
    )
    
    # Total AA = ITS_AA + CON_AA
    metrics['Total_AA'] = (
        metrics.get('ITS_AA_Total', 0) + 
        metrics.get('CON_AA_Total', 0)
    )
    
    # Manual Claims = Total Claims - Total AA
    metrics['Manual_Claims'] = metrics['Total_Claims'] - metrics['Total_AA']
    
    # Total Received = ITS_RCVD + CON_RCVD
    metrics['Total_RCVD'] = (
        metrics.get('ITS_RCVD_Total', 0) + 
        metrics.get('CON_RCVD_Total', 0)
    )
    
    # First Pass (System + Auto Reject + Recycle)
    first_pass_cols = [
        'ITS_SYS_AA_Total', 'ITS_AUTO_REJ_AA_Total', 'ITS_RECY_AA_Total',
        'CON_SYS_AA_Total', 'CON_AUTO_REJ_AA_Total', 'CON_RECY_AA_Total'
    ]
    metrics['First_Pass'] = sum(metrics.get(col, 0) for col in first_pass_cols)
    
    # Second Pass (OC + COGAI)
    second_pass_cols = [
        'ITS_OC_AA_Total', 'ITS_COGAI_AA_Total',
        'CON_OC_AA_Total', 'CON_COGAI_AA_Total'
    ]
    metrics['Second_Pass'] = sum(metrics.get(col, 0) for col in second_pass_cols)
    
    # Calculate AA Rate %
    if metrics['Total_Claims'] > 0:
        metrics['AA_Rate_Pct'] = round((metrics['Total_AA'] / metrics['Total_Claims']) * 100, 2)
    else:
        metrics['AA_Rate_Pct'] = 0.0
    
    if logger:
        logger.info(f"Metrics calculated - Total Claims: {metrics['Total_Claims']:,}, "
                   f"AA Rate: {metrics['AA_Rate_Pct']:.2f}%")
    
    return pd.DataFrame([metrics])


def update_excel_with_duplicate_check(excel_path, sheet_name, date_value, data_row, logger=None):
    """
    Update Excel workbook with new data, checking for duplicates first.
    Production pattern: Prevents duplicate data insertion by checking existing dates.
    
    Args:
        excel_path: Path to Excel file
        sheet_name: Name of worksheet to update
        date_value: Date value to check for duplicates (string format: 'MM/DD/YYYY')
        data_row: Pandas Series or dict with data to append
        logger: Logger instance
    
    Returns:
        bool: True if updated, False if duplicate found
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        if logger:
            logger.info(f"Updating Excel: {excel_path}, Sheet: {sheet_name}")
        
        # Load workbook
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook[sheet_name]
        
        # Read existing data to check for duplicates
        existing_df = pd.read_excel(excel_path, sheet_name=sheet_name)
        
        # Check if date already exists (assuming first column is date)
        date_column = existing_df.columns[0]
        if date_value in existing_df[date_column].values:
            if logger:
                logger.warning(f"Data for {date_value} already exists in Excel. Skipping update.")
            return False
        
        # Append new row
        last_row = worksheet.max_row + 1
        
        # Write date in first column
        worksheet.cell(row=last_row, column=1).value = date_value
        
        # Write data values
        if isinstance(data_row, pd.Series):
            for col_index, value in enumerate(data_row):
                # Convert numpy types to Python types for Excel compatibility
                if isinstance(value, (np.integer, np.floating)):
                    value = value.item()
                worksheet.cell(row=last_row, column=col_index + 2).value = value
        elif isinstance(data_row, dict):
            for col_index, (key, value) in enumerate(data_row.items()):
                if isinstance(value, (np.integer, np.floating)):
                    value = value.item()
                worksheet.cell(row=last_row, column=col_index + 2).value = value
        
        # Save workbook
        workbook.save(excel_path)
        
        if logger:
            logger.info(f"Successfully updated Excel with data for {date_value}")
        
        return True
        
    except FileNotFoundError:
        if logger:
            log_error_with_context(logger, FileNotFoundError(f"File not found: {excel_path}"), 
                                  "Excel Update")
        raise
    except PermissionError:
        if logger:
            log_error_with_context(logger, PermissionError(f"Permission denied: {excel_path}"), 
                                  "Excel Update - Please close the file")
        raise
    except Exception as e:
        if logger:
            log_error_with_context(logger, e, "Excel Update")
        raise


def format_dataframe_for_html(df, format_numbers=True):
    """
    Format DataFrame for HTML display with proper number formatting.
    Production pattern: Formats numbers with commas, preserves percentage formatting.
    
    Args:
        df: Input DataFrame
        format_numbers: Whether to format numeric columns with commas
    
    Returns:
        DataFrame with formatted values
    """
    df_formatted = df.copy()
    
    if format_numbers:
        # Get numeric columns
        numeric_cols = df_formatted.select_dtypes(include=['number']).columns
        
        for col in numeric_cols:
            # Format integers with commas, keep decimals for percentages
            col_lower = col.lower()
            if any(keyword in col_lower for keyword in ['rate', '%', 'pct', 'percent']):
                # Format as percentage with 2 decimals
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: f"{x:.2f}" if pd.notna(x) else ""
                )
            else:
                # Format as integer with commas
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: f"{int(x):,}" if pd.notna(x) and x == int(x) else f"{x:,.2f}" if pd.notna(x) else ""
                )
    
    return df_formatted


def generate_html_email_body(metrics_dict, report_date, title="Healthcare Claims Auto-Adjudication Report"):
    """
    Generate comprehensive HTML email body with multiple formatted tables.
    Production pattern: Creates styled HTML with multiple segment tables.
    
    Args:
        metrics_dict: Dictionary of DataFrames for each segment
                     Example: {'WGS': df_wgs, 'Medicaid': df_med, 'GBD': df_gbd}
        report_date: Date of the report (string)
        title: Email title/header
    
    Returns:
        str: HTML formatted email body
    """
    html_parts = [
        "<html>",
        "<head>",
        "<style>",
        "body { font-family: Arial, sans-serif; margin: 20px; }",
        "h2 { color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }",
        "h3 { color: #34495e; margin-top: 30px; }",
        "table { border-collapse: collapse; margin: 20px 0; width: 100%; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }",
        "th { background-color: #3498db; color: white; padding: 12px; text-align: left; font-weight: bold; }",
        "td { border: 1px solid #ddd; padding: 10px; text-align: right; }",
        "td:first-child { text-align: left; font-weight: bold; }",
        "tr:nth-child(even) { background-color: #f8f9fa; }",
        "tr:hover { background-color: #e8f4f8; }",
        "tr:last-child { background-color: #d5e8f3; font-weight: bold; }",
        ".footer { margin-top: 30px; padding: 15px; background-color: #ecf0f1; border-radius: 5px; }",
        "</style>",
        "</head>",
        "<body>",
        f"<h2>{title}</h2>",
        f"<p><strong>Report Date:</strong> {report_date}</p>",
        f"<p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
    ]
    
    # Add each segment's table
    for segment_name, df in metrics_dict.items():
        html_parts.append(f"<h3>{segment_name}</h3>")
        
        # Format DataFrame and convert to HTML
        df_formatted = format_dataframe_for_html(df)
        table_html = df_formatted.to_html(index=True, escape=False, border=0, classes='data-table')
        html_parts.append(table_html)
    
    # Add footer
    html_parts.extend([
        "<div class='footer'>",
        "<p><em>This is an automated report generated by the Healthcare Claims Reporting Pipeline.</em></p>",
        "<p><em>For questions or issues, please contact the Data Analytics team.</em></p>",
        "</div>",
        "</body>",
        "</html>"
    ])
    
    return "\n".join(html_parts)


def send_outlook_email(subject, html_body, to_list, cc_list=None, attachments=None, logger=None):
    """
    Send email via Outlook with HTML body and attachments.
    Production pattern: Uses win32com for Outlook automation with error handling.
    
    Args:
        subject: Email subject
        html_body: HTML formatted email body
        to_list: Semicolon-separated recipient emails or list of emails
        cc_list: Semicolon-separated CC emails or list of emails (optional)
        attachments: List of file paths to attach (optional)
        logger: Logger instance (optional)
    
    Returns:
        bool: True if sent successfully, False otherwise
    """
    # Check if email sending is enabled via environment variable
    send_email = os.getenv('SEND_EMAIL', 'false').lower() == 'true'
    
    if not send_email:
        if logger:
            logger.info("Email sending disabled (SEND_EMAIL environment variable not set to 'true')")
            logger.info(f"Would have sent email to: {to_list}")
        return False
    
    try:
        import win32com.client as win
        
        if logger:
            logger.info("Creating Outlook email...")
        
        # Create Outlook application instance
        outlook_app = win.Dispatch('Outlook.Application')
        mail = outlook_app.CreateItem(0)  # 0 = MailItem
        
        # Set email properties
        mail.Subject = subject
        mail.HTMLBody = html_body
        
        # Handle to_list (can be string or list)
        if isinstance(to_list, list):
            mail.To = '; '.join(to_list)
        else:
            mail.To = to_list
        
        # Handle cc_list
        if cc_list:
            if isinstance(cc_list, list):
                mail.CC = '; '.join(cc_list)
            else:
                mail.CC = cc_list
        
        # Add attachments
        if attachments:
            for attachment_path in attachments:
                if os.path.exists(attachment_path):
                    mail.Attachments.Add(os.path.abspath(attachment_path))
                    if logger:
                        logger.info(f"Attached: {attachment_path}")
                else:
                    if logger:
                        logger.warning(f"Attachment not found: {attachment_path}")
        
        # Send email
        mail.Send()
        
        if logger:
            logger.info(f"Email sent successfully to: {to_list}")
        
        return True
        
    except ImportError:
        if logger:
            logger.error("win32com.client not available. Install pywin32: pip install pywin32")
        return False
    except Exception as e:
        if logger:
            log_error_with_context(logger, e, "Email Sending")
        return False


def accumulate_ytd_data(new_data_df, ytd_file_path, date_column='ReportDate', logger=None):
    """
    Accumulate new data into YTD dataset with duplicate prevention.
    Production pattern: Appends to historical dataset while checking for duplicates.
    
    Args:
        new_data_df: DataFrame with new data to append
        ytd_file_path: Path to YTD CSV file
        date_column: Column name containing dates for duplicate checking
        logger: Logger instance
    
    Returns:
        DataFrame: Updated YTD dataset
    """
    try:
        if logger:
            log_section_start(logger, "YTD Data Accumulation")
        
        # Load existing YTD data if it exists
        if os.path.exists(ytd_file_path):
            ytd_df = pd.read_csv(ytd_file_path)
            if logger:
                logger.info(f"Loaded existing YTD data: {len(ytd_df)} records")
            
            # Check for duplicates based on date
            if date_column in new_data_df.columns and date_column in ytd_df.columns:
                new_dates = new_data_df[date_column].unique()
                existing_dates = ytd_df[date_column].unique()
                duplicate_dates = set(new_dates) & set(existing_dates)
                
                if duplicate_dates:
                    if logger:
                        logger.warning(f"Found duplicate dates: {duplicate_dates}")
                    # Remove duplicates from new data
                    new_data_df = new_data_df[~new_data_df[date_column].isin(duplicate_dates)]
                    if logger:
                        logger.info(f"Removed duplicates, {len(new_data_df)} new records to add")
            
            # Append new data
            ytd_df = pd.concat([ytd_df, new_data_df], ignore_index=True)
        else:
            if logger:
                logger.info("Creating new YTD dataset")
            ytd_df = new_data_df
        
        # Save updated YTD data
        ytd_df.to_csv(ytd_file_path, index=False)
        
        if logger:
            logger.info(f"YTD data updated: {len(ytd_df)} total records")
            log_section_end(logger, "YTD Data Accumulation", success=True)
        
        return ytd_df
        
    except Exception as e:
        if logger:
            log_error_with_context(logger, e, "YTD Data Accumulation")
            log_section_end(logger, "YTD Data Accumulation", success=False)
        raise


def create_grand_total_row(df, numeric_columns, rate_columns=None):
    """
    Create a grand total row for a DataFrame.
    Production pattern: Calculates totals and recalculates rates for grand total.
    
    Args:
        df: Input DataFrame
        numeric_columns: List of columns to sum
        rate_columns: List of rate columns to recalculate (optional)
    
    Returns:
        DataFrame with grand total row appended
    """
    df_with_total = df.copy()
    
    # Calculate grand total
    grand_total = {}
    for col in numeric_columns:
        if col in df.columns:
            grand_total[col] = df[col].sum()
    
    # Recalculate rates for grand total
    if rate_columns:
        for rate_col in rate_columns:
            if rate_col in df.columns:
                # Example: AA_Rate_Pct = (Total_AA / Total_Claims) * 100
                if 'Total_AA' in grand_total and 'Total_Claims' in grand_total:
                    if grand_total['Total_Claims'] > 0:
                        grand_total[rate_col] = round(
                            (grand_total['Total_AA'] / grand_total['Total_Claims']) * 100, 2
                        )
                    else:
                        grand_total[rate_col] = 0.0
    
    # Add grand total row
    df_with_total.loc['Grand Total'] = grand_total
    
    return df_with_total


# Example usage and testing
if __name__ == "__main__":
    print("Enhanced Pipeline Module - Production Patterns")
    print("=" * 60)
    print("\nAvailable functions:")
    print("1. process_by_segments() - Multi-segment data processing")
    print("2. calculate_advanced_metrics() - Comprehensive metric calculations")
    print("3. update_excel_with_duplicate_check() - Excel update with duplicate prevention")
    print("4. format_dataframe_for_html() - HTML table formatting")
    print("5. generate_html_email_body() - HTML email generation")
    print("6. send_outlook_email() - Outlook email automation")
    print("7. accumulate_ytd_data() - YTD data accumulation")
    print("8. create_grand_total_row() - Grand total calculation")
    print("\nImport this module in your main.py to use these functions.")

# Made with Bob
