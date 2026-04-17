# =======================
# REDACTED VERSION (for external sharing)
# Emails and local/network paths removed or parameterized.
# Email sending disabled unless SEND_EMAIL=true is set.
# =======================

# -*- coding: utf-8 -*- 
# -*- coding: utf-8 -*- 
""" 
""" 
import schedule 
import time 
import pytz 
import configparser 
from functools import total_ordering 
from importlib.resources import contents 
#from nis import match 
from operator import index 
import os 
from textwrap import indent 
import pandas as pd 
import datetime 
from dateutil.relativedelta import relativedelta 
import pyodbc 
import connect 
import aa_utilities1 
import pull_utilities1 
import csv 
from dateutil.relativedelta import relativedelta 
import calendar 
# import win32com.client as win  # REDACTED: disabled by default
SEND_EMAIL = os.getenv('SEND_EMAIL', 'false').lower() == 'true' 
import re 
def runAAReport(path, reportType, monthNum): 
 mbu_standard = ['GNC88612.txt', 'GNC88625.txt', 'GNC88637.txt', 'GNC88638.txt', 'GNC88639.txt', 'GNC88645.txt','GNC88651.txt', 'GNC88653.txt', 
 'GNC88512.txt', 'GNC88551.txt', 'GNC88553.txt'] 
 case_standard = ['GNC88632.txt', 'GNC88532.txt'] 
 mbu_hix = ['GNC88655.txt', 'GNC88554.txt'] 
 case_hix = ['GNC88656.txt'] 
 if reportType == 'mtd': 
 reports = {'GNC88512.txt' : 'AA_RATES_MBU'} 
 elif reportType == 'month': 
 reports = { 
 'GNC88612.txt' : 'AA_RATES_MBU', 
 } 
 for rep, db in reports.items(): 
 print(rep) 
 
 if rep in mbu_standard: 
 print("entered here") 
 df,test = aa_utilities1.parse612(rep) 
 global today 
 today=test 
 global updategncdate 
 updategncdate= today.replace("/","") 
 output_path7=r'C:\\REDACTED\\PATH'+updategncdate+'_(2024)'+'.csv' 
 csv_file_path = r'C:\\REDACTED\\PATH' 
 csv_df = pd.read_csv(csv_file_path) 
 df1 = pd.merge(df, csv_df, left_on='MBU', right_on='MBU', how='inner') 
 global pg_new_st 
 pg_new_st= df.copy() 
 pf1 = pd.merge(df, csv_df, left_on='MBU', right_on='MBU', how='inner') 
 netfile =r'C:\\REDACTED\\PATH' 
 mtddf1=pd.read_csv(netfile) 
 mtddf2=df1 
 mtdmerged=pd.concat([mtddf1,mtddf2],ignore_index=True) 
 mtdmerged.to_csv(output_path7) 
 l=['LOCAL - CA','LOCAL - NV','LOCAL - CO'] 
 lg=['LOCAL - CA','LOCAL - NV','LOCAL - CO'] 
 df1=df1[df1['LOB'].isin(l)] 
 la=['LOCAL - CA'] 
 lb=['LOCAL - NV'] 
 lc=['LOCAL - CO'] 
 dfa=df1[df1['LOB'].isin(la)] 
 dfb=df1[df1['LOB'].isin(lb)] 
 dfc=df1[df1['LOB'].isin(lc)] 
 selcols=['ITS_RCVD','ITS_FNLZD','ITS_AA','CON_RCVD','CON_FNLZD','CON_AA','TOT_CLMS','TOT_AA','ITS_SYS_AA','ITS_AUTO_REJ_AA','ITS_RECY_AA','CON_SYS_AA','CON_AUTO_REJ_AA','CON_RECY_AA','ITS_OC_AA','ITS_COGAI_AA','CON_OC_AA','CON_COGAI_AA'] 
 for i in dfa: 
  dfa[i]=pd.to_numeric(dfa[i],errors='coerce') 
 for i in dfb: 
  dfb[i]=pd.to_numeric(dfb[i],errors='coerce') 
 for i in dfc: 
  dfc[i]=pd.to_numeric(dfc[i],errors='coerce') 
 totdict1={} 
 for i in selcols: 
  totdict1[i+' Total']=dfa[i].sum() 
 totdict2={} 
 for i in selcols: 
  totdict2[i+' Total']=dfb[i].sum() 
 totdict3={} 
 for i in selcols: 
  totdict3[i+' Total']=dfc[i].sum() 
 Totaldf1=pd.DataFrame.from_dict(totdict1,orient='index').T 
 Totaldf2=pd.DataFrame.from_dict(totdict2,orient='index').T 
 Totaldf3=pd.DataFrame.from_dict(totdict3,orient='index').T 
 Totaldf1['Sum of TOT_AA']=Totaldf1['ITS_AA Total']+Totaldf1['CON_AA Total'] 
 Totaldf2['Sum of TOT_AA']=Totaldf2['ITS_AA Total']+Totaldf2['CON_AA Total'] 
 Totaldf3['Sum of TOT_AA']=Totaldf3['ITS_AA Total']+Totaldf3['CON_AA Total'] 
 Totaldf1['Sum of TOT_CLMS']=Totaldf1['ITS_FNLZD Total']+Totaldf1['CON_FNLZD Total'] 
 Totaldf2['Sum of TOT_CLMS']=Totaldf2['ITS_FNLZD Total']+Totaldf2['CON_FNLZD Total'] 
 Totaldf3['Sum of TOT_CLMS']=Totaldf3['ITS_FNLZD Total']+Totaldf3['CON_FNLZD Total'] 
 Totaldf1['Manual PROCSD']=Totaldf1['Sum of TOT_CLMS']-Totaldf1['Sum of TOT_AA'] 
 Totaldf2['Manual PROCSD']=Totaldf2['Sum of TOT_CLMS']-Totaldf2['Sum of TOT_AA'] 
 Totaldf3['Manual PROCSD']=Totaldf3['Sum of TOT_CLMS']-Totaldf3['Sum of TOT_AA'] 
 Totaldf1['Total RCVD']=Totaldf1['ITS_RCVD Total']+Totaldf1['CON_RCVD Total'] 
 Totaldf2['Total RCVD']=Totaldf2['ITS_RCVD Total']+Totaldf2['CON_RCVD Total'] 
 Totaldf3['Total RCVD']=Totaldf3['ITS_RCVD Total']+Totaldf3['CON_RCVD Total'] 
 Totaldf1['1st Pass']=Totaldf1[['ITS_SYS_AA Total','ITS_AUTO_REJ_AA Total','ITS_RECY_AA Total','CON_SYS_AA Total','CON_AUTO_REJ_AA Total','CON_RECY_AA Total']].sum(axis=1) 
 Totaldf2['1st Pass']=Totaldf2[['ITS_SYS_AA Total','ITS_AUTO_REJ_AA Total','ITS_RECY_AA Total','CON_SYS_AA Total','CON_AUTO_REJ_AA Total','CON_RECY_AA Total']].sum(axis=1) 
 Totaldf3['1st Pass']=Totaldf3[['ITS_SYS_AA Total','ITS_AUTO_REJ_AA Total','ITS_RECY_AA Total','CON_SYS_AA Total','CON_AUTO_REJ_AA Total','CON_RECY_AA Total']].sum(axis=1) 
 Totaldf1['2nd Pass']=Totaldf1[['ITS_OC_AA Total','ITS_COGAI_AA Total','CON_OC_AA Total','CON_COGAI_AA Total']].sum(axis=1) 
 Totaldf2['2nd Pass']=Totaldf2[['ITS_OC_AA Total','ITS_COGAI_AA Total','CON_OC_AA Total','CON_COGAI_AA Total']].sum(axis=1) 
 Totaldf3['2nd Pass']=Totaldf3[['ITS_OC_AA Total','ITS_COGAI_AA Total','CON_OC_AA Total','CON_COGAI_AA Total']].sum(axis=1) 
 Totaldf1['aa_rate %']=round(Totaldf1['Sum of TOT_AA']/Totaldf1['Sum of TOT_CLMS'],4)*100 
 Totaldf2['aa_rate %']=round(Totaldf2['Sum of TOT_AA']/Totaldf2['Sum of TOT_CLMS'],4)*100 
 Totaldf3['aa_rate %']=round(Totaldf3['Sum of TOT_AA']/Totaldf3['Sum of TOT_CLMS'],4)*100 
 newdf1=Totaldf1.iloc[:,18:25] 
 newdf2=Totaldf2.iloc[:,18:25] 
 newdf3=Totaldf3.iloc[:,18:25] 
 com=pd.concat([newdf1,newdf2,newdf3]) 
 com.index=lg 
 com.loc['Grand Total']= com.sum(numeric_only=True, axis=0) 
 com['aa_rate %']['Grand Total']=round(com['Sum of TOT_AA']['Grand Total']/com['Sum of TOT_CLMS']['Grand Total'],4)*100 
 intcols1=list(com)[:-1] 
 for c in intcols1: 
  com[c]=com[c].astype(int) 
 numeric_columns=com.select_dtypes(include=['number']).columns[:-1] 
 for colum in numeric_columns: 
  if pd.api.types.is_numeric_dtype(com[colum]): 
   com[colum]=com[colum].apply(lambda y: '{:,}'.format(y)) 
 htmltable=com.to_html() 
 import openpyxl 
 from openpyxl import load_workbook 
 workpath=r'C:\\REDACTED\\PATH' 
 workbook=openpyxl.load_workbook(workpath) 
 worksheet=workbook['Westmarketaarate'] 
 exceldf=pd.read_excel(workpath) 
 if today in exceldf['Unnamed: 0'].values: 
  print("This data already present in Excel") 
  break 
 else: 
  last_row=worksheet.max_row+1 
  worksheet.cell(row=last_row,column=1).value=today 
  for col_index,value in enumerate(com.iloc[-1]): 
   worksheet.cell(row=last_row,column=col_index+2).value=value 
  workbook.save(workpath) 
 px_gbd=pf1.copy() 
 px_com=pf1.copy() 
 global px_wgs 
 px_wgs = pf1.copy() 
 pg_new_st=pf1.copy() 
 pg=['SSB'] 
 pf1=pf1[pf1['BOB'].isin(pg)] 
 selcols1=['ITS_RCVD','ITS_FNLZD','ITS_AA','CON_RCVD','CON_FNLZD','CON_AA','TOT_CLMS','TOT_AA','ITS_SYS_AA','ITS_AUTO_REJ_AA','ITS_RECY_AA','CON_SYS_AA','CON_AUTO_REJ_AA','CON_RECY_AA','ITS_OC_AA','ITS_COGAI_AA','CON_OC_AA','CON_COGAI_AA'] 
 selcols2=['SEGMENT'] 
 pf2=pf1[selcols1] 
 pf3=pf1[selcols2] 
 for i in pf2: 
  pf2[i]=pd.to_numeric(pf2[i],errors='coerce') 
 pf4=pd.concat([pf2,pf3],axis=1) 
 finaldf=pf4.groupby(['SEGMENT']).sum() 
 newdf=finaldf[['TOT_AA','TOT_CLMS']] 
 newdf['Manual Claims']=finaldf['TOT_CLMS']-finaldf['TOT_AA'] 
 newdf['Total RCVD']=finaldf['ITS_RCVD']+finaldf['CON_RCVD'] 
 newdf['1st Pass']=finaldf[['ITS_SYS_AA','ITS_AUTO_REJ_AA','ITS_RECY_AA','CON_SYS_AA','CON_AUTO_REJ_AA','CON_RECY_AA']].sum(axis=1) 
 newdf['2nd Pass']=finaldf[['ITS_OC_AA','ITS_COGAI_AA','CON_OC_AA','CON_COGAI_AA']].sum(axis=1) 
 newdf['AA_Rate %']=round(finaldf['TOT_AA']/finaldf['TOT_CLMS'],4)*100 
 newdf.loc['Medicaid Total']= newdf.sum(numeric_only=True, axis=0) 
 for i in newdf.index: 
  print(i) 
 newdf['AA_Rate %']['Medicaid Total']=round(newdf['TOT_AA']['Medicaid Total']/newdf['TOT_CLMS']['Medicaid Total'],4)*100 
 intcols=list(newdf)[:-1] 
 for c in intcols: 
  newdf[c]=newdf[c].astype(int) 
 numeric_columns1=newdf.select_dtypes(include=['number']).columns[:-1] 
 for colum1 in numeric_columns1: 
  if pd.api.types.is_numeric_dtype(newdf[colum1]): 
   newdf[colum1]=newdf[colum1].apply(lambda y1: '{:,}'.format(y1)) 
 resfinal=newdf.reset_index() 
 htmltable1=resfinal.to_html(index=False) 
 workbook=openpyxl.load_workbook(workpath) 
 worksheet=workbook['Westmarketaarate'] 
 start='J' 
 start_col=openpyxl.utils.column_index_from_string(start) 
 start1='I' 
 start2=openpyxl.utils.column_index_from_string(start1) 
 b=worksheet.max_row 
 worksheet.cell(row=b,column=start2).value=today 
 for col_index,value in enumerate(newdf.iloc[-1]): 
  worksheet.cell(row=b,column=start_col+col_index).value=value 
 workbook.save(workpath) 
 pg_gbd=['SSB','SENIOR'] 
 px_gbd=px_gbd[px_gbd['BOB'].isin(pg_gbd)] 
 selcols1x=['ITS_RCVD','ITS_FNLZD','ITS_AA','CON_RCVD','CON_FNLZD','CON_AA','TOT_CLMS','TOT_AA','ITS_SYS_AA','ITS_AUTO_REJ_AA','ITS_RECY_AA','CON_SYS_AA','CON_AUTO_REJ_AA','CON_RECY_AA','ITS_OC_AA','ITS_COGAI_AA','CON_OC_AA','CON_COGAI_AA'] 
 selcols2x=['SEGMENT','BOB'] 
 pf2_gbd=px_gbd[selcols1x] 
 pf3_gbd=px_gbd[selcols2x] 
 for i in pf2_gbd: 
  pf2_gbd[i]=pd.to_numeric(pf2_gbd[i],errors='coerce') 
 pf4_gbd=pd.concat([pf2_gbd,pf3_gbd],axis=1) 
 finaldf_gbd=pf4_gbd.groupby(['BOB']).sum() 
 newdf_gbd=finaldf_gbd[['TOT_AA','TOT_CLMS']] 
 newdf_gbd['Manual Claims']=finaldf_gbd['TOT_CLMS']-finaldf_gbd['TOT_AA'] 
 newdf_gbd['Total RCVD']=finaldf_gbd['ITS_RCVD']+finaldf_gbd['CON_RCVD'] 
 newdf_gbd['1st Pass']=finaldf_gbd[['ITS_SYS_AA','ITS_AUTO_REJ_AA','ITS_RECY_AA','CON_SYS_AA','CON_AUTO_REJ_AA','CON_RECY_AA']].sum(axis=1) 
 newdf_gbd['2nd Pass']=finaldf_gbd[['ITS_OC_AA','ITS_COGAI_AA','CON_OC_AA','CON_COGAI_AA']].sum(axis=1) 
 newdf_gbd['AA_Rate %'] = ((finaldf_gbd['TOT_AA'] / finaldf_gbd['TOT_CLMS']) * 100).round(2) 
 newdf_gbd.loc['Grand_Total']= newdf_gbd.sum(numeric_only=True, axis=0) 
 newdf_gbd['AA_Rate %']['Grand_Total']=round(newdf_gbd['TOT_AA']['Grand_Total']/newdf_gbd['TOT_CLMS']['Grand_Total'],4)*100 
 gbd_final=newdf_gbd 
 grand_total_gbd_final =gbd_final.applymap(lambda x: f"{int(x):,}" if isinstance(x, (int, float)) and x == int(x) else f"{x:,.2f}" if isinstance(x, float) else x) 
 grand_total_gbd_final.index=["SENIOR","SSB","GBD_Rates"] 
 gbd_rates_only = grand_total_gbd_final.loc[['GBD_Rates']] 
 htmltable_gbd = gbd_rates_only.to_html(index=True) 
 pg_gbd = ['SSB', 'SENIOR'] 
 px_com = px_com[~px_com['BOB'].isin(pg_gbd)] 
 selcols1x_com = ['ITS_RCVD', 'ITS_FNLZD', 'ITS_AA', 'CON_RCVD', 'CON_FNLZD', 'CON_AA', 'TOT_CLMS', 'TOT_AA', 'ITS_SYS_AA', 'ITS_AUTO_REJ_AA', 'ITS_RECY_AA', 'CON_SYS_AA', 'CON_AUTO_REJ_AA', 'CON_RECY_AA', 'ITS_OC_AA', 'ITS_COGAI_AA', 'CON_OC_AA', 'CON_COGAI_AA'] 
 pf2_com = px_com[selcols1x_com] 
 for i in pf2_com: 
  pf2_com[i] = pd.to_numeric(pf2_com[i], errors='coerce') 
 pf4_com=pf2_com 
 grand_total = pf4_com.sum() 
 pf4_com.loc['Grand Total'] = grand_total 
 newdf_com = pd.DataFrame() 
 newdf_com=pf4_com[['TOT_AA','TOT_CLMS']] 
 newdf_com['Manual Claims'] = pf4_com['TOT_CLMS'] - pf4_com['TOT_AA'] 
 newdf_com['Total RCVD'] = pf4_com['ITS_RCVD'] + pf4_com['CON_RCVD'] 
 newdf_com['1st Pass'] = pf4_com[['ITS_SYS_AA', 'ITS_AUTO_REJ_AA', 'ITS_RECY_AA', 'CON_SYS_AA', 'CON_AUTO_REJ_AA', 'CON_RECY_AA']].sum(axis=1) 
 newdf_com['2nd Pass'] = pf4_com[['ITS_OC_AA', 'ITS_COGAI_AA', 'CON_OC_AA', 'CON_COGAI_AA']].sum(axis=1) 
 newdf_com['AA_Rate %'] = round(pf4_com['TOT_AA'] / pf4_com['TOT_CLMS'], 4) * 100 
 grand_total_df = newdf_com.tail(1) 
 grand_total_df_formatted = grand_total_df.applymap(lambda x: f"{x:,}" if isinstance(x, (int, float)) else x) 
 htmltable_commercial=grand_total_df_formatted.to_html(index=False) 
 pg_wgs = ['SSB', 'SENIOR'] 
 px_wgs['BOB'] = px_wgs['BOB'].apply(lambda x: x if x in pg_wgs else 'COMMERICAL') 
 selcols1x = ['ITS_RCVD', 'ITS_FNLZD', 'ITS_AA', 'CON_RCVD', 'CON_FNLZD', 'CON_AA', 'TOT_CLMS', 'TOT_AA', 'ITS_SYS_AA', 'ITS_AUTO_REJ_AA', 'ITS_RECY_AA', 'CON_SYS_AA', 'CON_AUTO_REJ_AA', 'CON_RECY_AA', 'ITS_OC_AA', 'ITS_COGAI_AA', 'CON_OC_AA', 'CON_COGAI_AA'] 
 selcols2x = ['SEGMENT', 'BOB'] 
 pf2_wgs = px_wgs[selcols1x] 
 global pf3_wgs 
 pf3_wgs = px_wgs[['BOB']] 
 for i in pf2_wgs: 
  pf2_wgs[i] = pd.to_numeric(pf2_wgs[i], errors='coerce') 
 pf4_wgs = pd.concat([pf2_wgs, pf3_wgs], axis=1) 
 finaldf_wgs = pf4_wgs.groupby(['BOB']).sum() 
 newdf_wgs = finaldf_wgs[['TOT_AA', 'TOT_CLMS']] 
 newdf_wgs['Manual Claims'] = finaldf_wgs['TOT_CLMS'] - finaldf_wgs['TOT_AA'] 
 newdf_wgs['Total RCVD'] = finaldf_wgs['ITS_RCVD'] + finaldf_wgs['CON_RCVD'] 
 newdf_wgs['1st Pass'] = finaldf_wgs[['ITS_SYS_AA', 'ITS_AUTO_REJ_AA', 'ITS_RECY_AA', 'CON_SYS_AA', 'CON_AUTO_REJ_AA', 'CON_RECY_AA']].sum(axis=1) 
 newdf_wgs['2nd Pass'] = finaldf_wgs[['ITS_OC_AA', 'ITS_COGAI_AA', 'CON_OC_AA', 'CON_COGAI_AA']].sum(axis=1) 
 newdf_wgs['AA_Rate %'] = ((finaldf_wgs['TOT_AA'] / finaldf_wgs['TOT_CLMS']) * 100).round(2) 
 newdf_wgs.loc['Grand_Total'] = newdf_wgs.sum(numeric_only=True, axis=0) 
 newdf_wgs['AA_Rate %']['Grand_Total'] = round(newdf_wgs['TOT_AA']['Grand_Total'] / newdf_wgs['TOT_CLMS']['Grand_Total'], 4) * 100 
 wgs_final = newdf_wgs 
 grand_total_wgs_final = wgs_final.applymap(lambda x: f"{int(x):,}" if isinstance(x, (int, float)) and x == int(x) else f"{x:,}") 
 desired_order = ["COMMERICAL", "SSB", "SENIOR", "Grand_Total"] 
 grand_total_wgs_final = grand_total_wgs_final.loc[desired_order] 
 if len(grand_total_wgs_final) == 4: 
  grand_total_wgs_final.index = ["COMMERICAL", "SSB", "SENIOR", "WGS_RATES"] 
 else: 
  print("Error: The number of rows in the DataFrame does not match the expected number of index labels.") 
 htmltable_wgs = grand_total_wgs_final.to_html(index=True) 
 regions = ['LOCAL - MD - SG', 'LOCAL - FL - SG', 'LOCAL - TX - SG'] 
 pg_new_st = pg_new_st['SEGMENT'].isin(regions) 
 selcols1x_st = ['SEGMENT','ITS_RCVD', 'ITS_FNLZD', 'ITS_AA', 'CON_RCVD', 'CON_FNLZD', 'CON_AA', 'TOT_CLMS', 'TOT_AA', 'ITS_SYS_AA', 'ITS_AUTO_REJ_AA', 'ITS_RECY_AA', 'CON_SYS_AA', 'CON_AUTO_REJ_AA', 'CON_RECY_AA', 'ITS_OC_AA', 'ITS_COGAI_AA', 'CON_OC_AA', 'CON_COGAI_AA'] 
 pf4_new_st = pg_new_st[selcols1x_st] 
 global finaldf_new_st 
 for col in selcols1x_st: 
  if col != 'SEGMENT': 
   pf4_new_st[col] = pd.to_numeric(pf4_new_st[col], errors='coerce') 
 finaldf_new_st = pf4_new_st.groupby(['SEGMENT']).sum() 
 newdf_new_st = finaldf_new_st[['TOT_AA', 'TOT_CLMS']] 
 newdf_new_st['Manual Claims'] = finaldf_new_st['TOT_CLMS'] - finaldf_new_st['TOT_AA'] 
 newdf_new_st['Total RCVD'] = finaldf_new_st['ITS_RCVD'] + finaldf_new_st['CON_RCVD'] 
 newdf_new_st['1st Pass'] = finaldf_new_st[['ITS_SYS_AA', 'ITS_AUTO_REJ_AA', 'ITS_RECY_AA', 'CON_SYS_AA', 'CON_AUTO_REJ_AA', 'CON_RECY_AA']].sum(axis=1) 
 newdf_new_st['2nd Pass'] = finaldf_new_st[['ITS_OC_AA', 'ITS_COGAI_AA', 'CON_OC_AA', 'CON_COGAI_AA']].sum(axis=1) 
 newdf_new_st['AA_Rate %'] = ((finaldf_new_st['TOT_AA'] / finaldf_new_st['TOT_CLMS']) * 100).round(2) 
 new_st_final = newdf_new_st 
 new_st_final.reset_index(inplace=True) 
 grand_total_new_st_final = new_st_final.applymap(lambda x: f"{int(x):,}" if isinstance(x, (int, float)) and x == int(x) else f"{x:,.2f}" if isinstance(x, (int, float)) else x) 
 grand_total_new_st_final['SEGMENT'] = grand_total_new_st_final['SEGMENT'].replace({ 'LOCAL - FL - SG': 'FL', 'LOCAL - MD - SG': 'MD', 'LOCAL - TX - SG': 'TX' }) 
 htmltable_new_st = grand_total_new_st_final.to_html(index=False) 
 TO_LIST = os.getenv('TO_LIST', '')
 CC_LIST = os.getenv('CC_LIST', '') 
 SENDER_MAIL = os.getenv('SENDER_MAIL', '')
 st=str(today) 
 sub='ALL WGS Report : Month To Date Report - '+st 
 if SEND_EMAIL:
    import win32com.client as win
    outlook_app=win.Dispatch('Outlook.Application') 
     mail=outlook_app.CreateItem(0) 
     mail.Subject=sub 
 html_body = f"""<html><body>...{htmltable_wgs}{htmltable_gbd}{htmltable_new_st}{htmltable1}{htmltable}</body></html>""" 
     mail.HTMLBody = html_body 
     mail.To = TO_LIST  # was: to_rec 
     mail.CC = CC_LIST  # was: cc_rec 
     attachment = output_path7 
     att = workpath 
     try:
        mail.Attachments.Add(attachment)
        
    except Exception as e:
        print('Attachment add skipped:', e) 
  
     mail.Send() 
     print('mail sent successfully')
else:
    print('SEND_EMAIL is False — email creation/sending disabled in redacted script.') 
if __name__ == '__main__': 
 config = configparser.ConfigParser() 
 code_path = connect.returnPath() 
 config_file = code_path + r"\\config.ini" 
 config.read(config_file) 
 dbname = config['DEFAULT']['DBNAME'] 
 dbpath = config['DEFAULT']['DBPATH'] 
 path = config['DEFAULT']['ACCESS'] 
 os.chdir(path) 
 current_date = datetime.datetime.today() 
 clm_system = 'WGS' 
 if current_date.day >=15 or current_date.day <=15: 
  print('midmonth') 
  month = current_date 
  month_num = month.month 
  month_txt = month.strftime("%B") 
  runAAReport(path, 'mtd', month_num) 
